import pandas as pd
import openpyxl
from openpyxl.styles import Font
import os

# --- CONFIGURATION ---
BASE_PATH = "/home/mfresq/PycharmProjects/marguerite/decidon/data/mandats/"
INPUT_FILE = os.path.join(BASE_PATH, "full_table_final_all_links.xlsx")
OUTPUT_FILE = os.path.join(BASE_PATH, "DECIDON_Final_Total_Links.xlsx")


def run_final_processor():
    print(f"Chargement de {INPUT_FILE}...")
    if not os.path.exists(INPUT_FILE):
        print("❌ Erreur : Fichier introuvable.")
        return

    # 1. Lire les données avec Pandas
    df = pd.read_excel(INPUT_FILE)

    # 2. Sauvegarder temporairement en Excel standard pour obtenir la structure
    df.to_excel(OUTPUT_FILE, index=False)

    # 3. Ouvrir avec Openpyxl pour injecter les vrais liens
    print("Injection des liens natifs (méthode robuste)...")
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active

    # Trouver les indices des colonnes (Openpyxl commence à 1)
    headers = [cell.value for cell in ws[1]]

    col_map = {
        'person_wikidata_qid': 'https://www.wikidata.org/wiki/',
        'person_wikipedia_url': '',  # URL directe
        'person_sycomore_id': 'https://www2.assemblee-nationale.fr/sycomore/fiche/(num_dept)/',
        'person_senat_id': 'https://www.senat.fr/senateur-3eme-republique/',
        'mandature_wikidata_qid': 'https://www.wikidata.org/wiki/',
        'mandature_wikipedia_url': ''  # URL directe
    }

    # Style pour les liens
    link_font = Font(color="0000FF", underline="single")

    # Parcourir les colonnes cibles
    for col_name, base_url in col_map.items():
        if col_name in headers:
            col_idx = headers.index(col_name) + 1
            print(f"  -> Traitement de la colonne : {col_name}")

            # Parcourir chaque ligne (en commençant par la ligne 2 pour éviter l'en-tête)
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = str(cell.value).strip()

                if val and val.lower() != 'nan':
                    target_url = ""

                    if col_name == 'person_sycomore_id':
                        target_url = f"{base_url}{val.replace('.0', '')}"
                    elif col_name == 'person_senat_id':
                        target_url = f"{base_url}{val}.html"
                    elif base_url == '':  # Pour les colonnes Wikipedia déjà en URL
                        target_url = val
                    else:  # Pour Wikidata
                        target_url = f"{base_url}{val}"

                    # Application du lien et du style
                    if target_url.startswith('http'):
                        cell.hyperlink = target_url
                        cell.font = link_font

    # 4. Sauvegarde finale
    print("Sauvegarde du fichier final...")
    wb.save(OUTPUT_FILE)
    print(f"✅ TERMINÉ ! Le fichier est prêt : {OUTPUT_FILE}")


if __name__ == "__main__":
    run_final_processor()